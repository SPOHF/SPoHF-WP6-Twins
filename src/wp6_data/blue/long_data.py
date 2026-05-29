"""Blue ``long_data`` source — yearly manual blueberry measurements.

``long_data`` is delivered as one Excel file per year (``Long_Data <year>.xlsx``)
in long/tidy form: ``Date, Meting, Treatment, Value`` (+ ``Plant_nr`` from 2025
on). This module is the only ``long_data``-specific code: it harmonizes the two
yearly vocabularies into one canonical set of measures and treatments, models
**one device per treatment**, and preserves every sample. The ``Plant_nr``
column is tolerated on input but discarded — plants are not individually
identified (see ADR 0004); every plant and pooled sample of a treatment coexists
on its single device.

It does **not** use the shared mean-bucketing ``bind``: because the source is
date-only, the unused time-of-day encodes each sample's file order
(``date 00:00:00 UTC + i seconds``), so many samples coexist at one
device+date+measure without averaging and without per-sample devices
(see ADR 0001). Ingest replaces only the calendar year(s) in the uploaded file
(see ADR 0002). Treatment codes reuse the automated sensors' ``position``
vocabulary so manual and automated readings of a plot group together.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import UTC, date, datetime, timedelta
from io import BytesIO

from openpyxl import load_workbook

from wp6_data.shared.manual_ingest import (
    ManualParseError,
    ManualSource,
    Reading,
    SkippedRow,
    ValidationReport,
)
from wp6_data.shared.manual_ingest.parsing import build_report

# Categorical value written to blue's manual-ingest `readings` column.
SOURCE = "long_data"

# Source treatment labels (2024 English, 2025 Dutch/codes) → canonical code,
# reused from the automated sensors' `position` vocabulary. The four 2025
# phenological-stage regimes pass through unchanged (V_=vegetative stage,
# G_=generative stage); they are first-class treatments, not Ca/K variants.
TREATMENT_MAP: dict[str, str] = {
    "Organic 1": "Org1", "Organisch-1": "Org1",
    "Organic 2": "Org2", "Organisch-2": "Org2",
    "Standard": "Std", "Standaard": "Std",
    "Calcium": "Ca", "Ca": "Ca",
    "Kalium": "K", "K": "K",
    "G_K": "G_K",
    "V_CA": "V_CA",
    "V_CA_G_BrPK": "V_CA_G_BrPK",
    "V_K_G_CaBrP": "V_K_G_CaBrP",
}

# Source measure labels (`Meting`) → canonical sensor_tag. Fresh vs
# after-storage are distinct tags; 2025's single Brix/Firmness map to the fresh
# tag; 2024 before-storage berry weight and 2025 per-berry weight merge into
# `berry_weight`; the Dutch `Oogst gewicht per plant` aliases `yield_per_plant`.
MEASURE_MAP: dict[str, str] = {
    "Shoot_Length": "shoot_length",
    "Brix Sugar content (pre storage)": "brix",
    "Brix": "brix",
    "Brix Sugar content (after storage)": "brix_after_storage",
    "Firmness (pre storing)": "firmness",
    "Firmness": "firmness",
    "Firmness (after storing)": "firmness_after_storage",
    "Weight per Berry before storing": "berry_weight",
    "Weigth per berry": "berry_weight",
    "Weight per berry after storing": "berry_weight_after_storage",
    "Weigth before storage": "sample_weight_before_storage",
    "Weigth after storage": "sample_weight_after_storage",
    "Yield per plant": "yield_per_plant",
    "Oogst gewicht per plant": "yield_per_plant",
    "Score": "score",
    "Budscore": "budscore",
    "Flowering score": "flowering_score",
}

# Cumulative/derived measures we deliberately do not store (silently dropped,
# not reported as skipped errors).
IGNORE_MEASURES: frozenset[str] = frozenset({
    "Total Yield per plant",
    "Total Weight (grams)",
})

# Required source columns (matched case-insensitively). A `Plant_nr` column may
# also be present (2025+); it is read past and discarded — see ADR 0004.
_REQUIRED_COLUMNS = ("date", "meting", "treatment", "value")


class LongDataParseError(ManualParseError):
    """The Excel file is structurally unparseable (missing required column).

    Per-row failures (bad date/value, unknown treatment/measure) are reported
    as skipped rows; only a missing required column fast-fails so the route
    factory renders a friendly rejection page.
    """


def _column_index(header: tuple) -> dict[str, int]:
    """Map column role → index from the header row (case-insensitive)."""
    found: dict[str, int] = {}
    for idx, cell in enumerate(header):
        if cell is None:
            continue
        name = str(cell).strip().lower()
        if name in _REQUIRED_COLUMNS:
            found[name] = idx
    missing = [c for c in _REQUIRED_COLUMNS if c not in found]
    if missing:
        raise LongDataParseError(
            f"missing required column(s): {', '.join(missing)}"
        )
    return found


def _coerce_date(cell: object) -> date | None:
    """Coerce an Excel date cell to a ``date`` (``None`` if unparseable)."""
    if isinstance(cell, datetime):
        return cell.date()
    if isinstance(cell, date):
        return cell
    if isinstance(cell, str):
        try:
            return date.fromisoformat(cell.strip()[:10])
        except ValueError:
            return None
    return None


def _aggregate(file_bytes: bytes) -> tuple[list[Reading], list[SkippedRow], int]:
    """Decode the workbook into ordinal-timestamped readings.

    Returns ``(readings, skipped, total)`` where ``total`` counts only rows we
    attempted to store (kept or skipped) — blank and ignored-measure rows are
    dropped as filler, so ``valid_rows == len(readings)``.
    """
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise LongDataParseError("file is empty") from None
    cols = _column_index(tuple(header))

    # Intermediate samples in file order; ordinals assigned after the full pass
    # so each (device, date, sensor) group is numbered by entry sequence.
    samples: list[tuple[str, date, str, float]] = []
    skipped: list[SkippedRow] = []
    total = 0

    for line_idx, row in enumerate(rows, start=2):
        if row is None or all(c is None for c in row):
            continue  # blank line — not a data row

        meting = row[cols["meting"]]
        meting = str(meting).strip() if meting is not None else ""
        if meting in IGNORE_MEASURES:
            continue  # cumulative/derived — deliberately dropped
        sensor_tag = MEASURE_MAP.get(meting)
        if sensor_tag is None:
            total += 1
            skipped.append(SkippedRow(line_idx, f"unrecognised measure {meting!r}"))
            continue

        treatment = row[cols["treatment"]]
        treatment = str(treatment).strip() if treatment is not None else ""
        code = TREATMENT_MAP.get(treatment)
        if code is None:
            total += 1
            skipped.append(
                SkippedRow(line_idx, f"unknown treatment {treatment!r}"),
            )
            continue

        day = _coerce_date(row[cols["date"]])
        if day is None:
            total += 1
            skipped.append(SkippedRow(line_idx, "date is missing or unparseable"))
            continue

        raw_value = row[cols["value"]]
        if raw_value is None or (isinstance(raw_value, str) and raw_value.strip() == ""):
            continue  # blank cell → measurement absent, not an error
        try:
            value = float(raw_value)
        except (TypeError, ValueError):
            total += 1
            skipped.append(SkippedRow(line_idx, f"value {raw_value!r} is not a number"))
            continue

        # The treatment code is the device — every plant and pooled sample of a
        # treatment shares one device (ADR 0004); any Plant_nr column is ignored.
        total += 1
        samples.append((code, day, sensor_tag, value))

    wb.close()

    # Assign the per-(device, date, sensor) sample ordinal: sample i → midnight
    # UTC + i seconds (i ≥ 1, so 00:00:00 stays reserved). Preserves spread and
    # entry order; UTC anchor keeps the calendar date correct in both
    # day-bucketing paths.
    ordinal: dict[tuple[str, date, str], int] = defaultdict(int)
    readings: list[Reading] = []
    for device_name, day, sensor_tag, value in samples:
        key = (device_name, day, sensor_tag)
        ordinal[key] += 1
        ts = datetime(day.year, day.month, day.day, tzinfo=UTC) + timedelta(
            seconds=ordinal[key],
        )
        readings.append(
            Reading(
                source=SOURCE,
                device_name=device_name,
                sensor_tag=sensor_tag,
                time=ts,
                value=value,
            ),
        )
    return readings, skipped, total


def parse(file_bytes: bytes) -> list[Reading]:
    return _aggregate(file_bytes)[0]


def validate(file_bytes: bytes) -> ValidationReport:
    readings, skipped, total = _aggregate(file_bytes)
    return build_report(file_bytes, readings, skipped, total)


def _calendar_year_scope(date_range: tuple[date, date]) -> tuple[str, list]:
    """Restrict replace/compare to the whole calendar year(s) in the upload, so
    one ``long_data`` source fed by yearly files replaces only that year."""
    lo, hi = date_range
    start = date(lo.year, 1, 1)
    end = date(hi.year + 1, 1, 1)
    return " AND time >= %s AND time < %s", [start, end]


LONG_DATA = ManualSource(
    slug="long_data",
    categorical_value=SOURCE,
    display_name="Long-form measurements",
    file_suffix=".xlsx",
    accept=".xlsx,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    row_noun="Excel rows",
    upload_hint=(
        "Upload a yearly Long_Data .xlsx file "
        "(columns: Date, Meting, Treatment, Value[, Plant_nr])."
    ),
    parse=parse,
    validate=validate,
    parse_error=LongDataParseError,
    replace_scope=_calendar_year_scope,
)
