"""Parser for the Sijia (Neurath) seasonal greenhouse Excel dataset.

This is the source-specific half of the manual-ingest capability: the
Excel/openpyxl reading, the Sijia column→sensor map, the Neurath device
naming, and the percentage scaling. The twin-agnostic data types
(``Reading``/``SkippedRow``/``ValidationReport``) and the base parse error
now live in ``wp6_data.shared.manual_ingest`` and are re-exported here so the
existing Sijia parser test suite and importers are unaffected.
"""

import io
from collections.abc import Iterator
from datetime import datetime, time

import openpyxl

from wp6_data.shared.manual_ingest.parsing import DecodedRow, bind
from wp6_data.shared.manual_ingest.types import (
    ManualParseError,
    Reading,
    SkippedRow,
    ValidationReport,
)

__all__ = [
    "COLUMN_TO_SENSOR",
    "DEFAULT_MEASUREMENT_HOUR",
    "EXPECTED_HEADERS",
    "META_COLUMNS",
    "PERCENTAGE_SENSORS",
    "SHEET_NAME",
    "SOURCE",
    "Reading",
    "SijiaParseError",
    "SkippedRow",
    "ValidationReport",
    "parse",
    "validate",
]

SOURCE = "sijia"
SHEET_NAME = "2025-26_Measurement"

# Excel rows record only a date, no time of day. Anchor measurements at 13:00
# so charts plot them at a sensible mid-day point (and so all readings from a
# single visit cluster at the same instant per (device, sensor)).
DEFAULT_MEASUREMENT_HOUR = 13

META_COLUMNS: tuple[str, ...] = ("Sample No.", "Variety", "Block", "Row", "Date")

COLUMN_TO_SENSOR: dict[str, str] = {
    "ChlM": "chlorophyll",
    "FlvM": "flavonoids",
    "AnthM": "anthocyanins",
    "NFI": "nfi",
    "Water Content (% of weight)": "water_content",
    "Minerals (% of weight)": "minerals",
    "Size Ø (mm)": "diameter",
    "Weight (g)": "weight",
    "Vitamin C – Fresh Sample (mg/100 g)": "vitamin_c_fresh",
    "Vitamin C – Dry Matter (mg/100 g)": "vitamin_c_dry",
    "Total Phenols – Fresh Sample (mg GAE/100 g)": "total_phenols_fresh",
    "Total Phenols – Dry Matter (mg GAE/100 g)": "total_phenols_dry",
    "Antioxidant Capacity – Fresh Sample (mmol TAE/100 g)": "antioxidant_capacity_fresh",
    "Antioxidant Capacity – Dry Matter (mmol TAE/100 g)": "antioxidant_capacity_dry",
}

# Headers compared exactly (em-dash sensitive). Order matters.
EXPECTED_HEADERS: tuple[str, ...] = META_COLUMNS + tuple(COLUMN_TO_SENSOR.keys())

# Sensors whose Excel value is a 0..1 fraction; parser scales to percentage
# so Y-axes are consistent with sensor data conventions (e.g. humidity %RH).
PERCENTAGE_SENSORS: frozenset[str] = frozenset({"water_content", "minerals"})


class SijiaParseError(ManualParseError):
    """Sijia file is structurally unparseable (wrong sheet, wrong headers).

    Per-row dtype failures are reported via ValidationReport.skipped_rows
    rather than raised — only structural failures fast-fail. Subclasses the
    shared ManualParseError so the shared route factory can render a friendly
    rejection page for it.
    """


def _device_name(variety: str, block: str, row: float) -> str:
    return f"neurath-{block}-{int(row)}-{variety.strip().lower()}"


_DATE_COL_IDX = EXPECTED_HEADERS.index("Date")
_VARIETY_COL_IDX = EXPECTED_HEADERS.index("Variety")
_BLOCK_COL_IDX = EXPECTED_HEADERS.index("Block")
_ROW_COL_IDX = EXPECTED_HEADERS.index("Row")
_SENSOR_COL_INDICES: tuple[tuple[int, str], ...] = tuple(
    (EXPECTED_HEADERS.index(col), sensor)
    for col, sensor in COLUMN_TO_SENSOR.items()
)


def _decode(file_bytes: bytes) -> Iterator[DecodedRow | SkippedRow]:
    """Decode the Sijia workbook into one row per measurement.

    The only Sijia-specific code: structural validation, the Neurath device
    naming, the date→13:00 anchoring and the percentage scaling. Bucketing,
    mean-aggregation and the ValidationReport are the shared scaffold.

    Implementation note: openpyxl in read_only mode is used directly rather
    than pandas read_excel because the Sijia file routinely carries 6+
    extra sheets (Klima, Pivot, Diagram) that bloat the workbook to 15 MB+.
    pandas would parse all of them and iterate the worksheet's full padded
    1,048,576-row dimension; openpyxl's streaming reader plus an early
    break on a NULL Date column reads only the actual ~120 measurement rows.
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes), read_only=True, data_only=True,
    )
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise SijiaParseError(
                f"Required sheet {SHEET_NAME!r} not found; "
                f"file contains {wb.sheetnames!r}"
            )
        ws = wb[SHEET_NAME]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise SijiaParseError(f"Sheet {SHEET_NAME!r} is empty") from None

        actual_headers = tuple(header_row[: len(EXPECTED_HEADERS)])
        if actual_headers != EXPECTED_HEADERS:
            raise SijiaParseError(
                f"Column headers mismatch. Expected {EXPECTED_HEADERS!r}, "
                f"got {actual_headers!r}"
            )

        for excel_idx, raw_row in enumerate(rows_iter, start=2):
            # First all-empty / NULL-Date row marks end of measurements;
            # Excel pads worksheets to 1,048,576 rows that we must NOT iterate.
            if raw_row[_DATE_COL_IDX] is None:
                break

            variety = raw_row[_VARIETY_COL_IDX]
            block = raw_row[_BLOCK_COL_IDX]
            row_num = raw_row[_ROW_COL_IDX]
            device = _device_name(variety, block, row_num)

            ts = raw_row[_DATE_COL_IDX]
            if not isinstance(ts, datetime):
                ts = datetime.combine(ts, time(0, 0))
            if ts.time() == time(0, 0):
                ts = ts.replace(hour=DEFAULT_MEASUREMENT_HOUR)

            row_cells: list[tuple[str, float]] = []
            error: str | None = None
            for col_idx, sensor in _SENSOR_COL_INDICES:
                value = raw_row[col_idx]
                if value is None:
                    continue
                try:
                    scaled = (
                        float(value) * 100 if sensor in PERCENTAGE_SENSORS
                        else float(value)
                    )
                except (TypeError, ValueError):
                    error = f"{EXPECTED_HEADERS[col_idx]}: {value!r} is not a number"
                    break
                row_cells.append((sensor, scaled))

            if error is not None:
                yield SkippedRow(row_index=excel_idx, reason=error)
            else:
                yield DecodedRow(
                    device_name=device, time=ts, cells=tuple(row_cells),
                )
    finally:
        wb.close()


parse, validate = bind(SOURCE, _decode)
